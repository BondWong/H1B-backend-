#!/usr/bin/env python3
import sys
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor
from openpyxl.cell import get_column_letter
import requests
import json
from requester import post
import threading
import queue

class ProducerThread(threading.Thread):
    def __init__(self):
        super(ProducerThread, self).__init__()

    def run(self):
        data = []
        i = 2
        for row in ws.iter_rows(row_offset = 2):
            d = {"name": "", "contact": "", "state": "", "city": "", "zipCode": "", "street": "", "title": "", "salary": "", "visaType": "H-1B", "workState": "", "workCity":"",  "workZipCode": ""}
            i += 1
            for cell in row:
                if not cell.value:
                    continue

                col = get_column_letter(cell.column)
                if col == "H":
                    d["name"] = cell.value
                elif col == "I":
                    d["street"] = cell.value
                elif col == "K":
                    d["city"] = cell.value
                elif col == "L":
                    d["state"] = cell.value
                elif col == "M":
                    d["zipCode"] = cell.value
                elif col == "P":
                    d["contact"] = cell.value
                elif col == "U":
                    d["title"] = cell.value
                elif col == "AA":
                    d["salary"] = "$" + cell.value
                elif col == "AB":
                    d["salary"] += "/" + cell.value
                elif col == "AK":
                    d["workCity"] = cell.value
                elif col == "AM":
                    d["workState"] = cell.value
                elif col == "AN":
                    d["workZipCode"] = cell.value
                elif col == "E" and cell.value != "H-1B":
                    continue;
            
                data.append(d)
                if len(data) >= 1000:
                    while queue.full():
                        pass
                    queue.put(data)
                    data = []

if __name__ == "__main__":
    filePath = sys.argv[1]
    url = sys.argv[2]
    tasks = {}
    wb = load_workbook(filePath, read_only=True)
    ws = wb.worksheets[0]
    queue = queue.Queue(50)

    producer = ProducerThread()
    producer.start()

    with ThreadPoolExecutor(32) as pool:
        while(True):
            if(not queue.empty()):
                data = queue.get()
                pool.submit(post, url, data)

    
